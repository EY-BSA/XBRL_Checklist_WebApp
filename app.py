"""
app.py  — XBRL 체크리스트 웹 애플리케이션 (Flask)
입력: IxD 편집기 '구조내려받기' .xlsx 파일 전용 
"""
import io, os, re, json, zipfile, traceback
from flask import Flask, request, jsonify, render_template, send_file

from taxonomy_xlsx_parser import parse_taxonomy_xlsx
from checklist_engine import run_all_checks, get_summary
from standard_taxonomy import (StandardTaxonomy,
                               enrich_axis_domain_check, enrich_dart_negate_check)

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024   # 100 MB

# 단순 메모리 캐시 (프로세스 내 마지막 결과 보관)
_cache: dict = {}

# 데이터 파일 경로
_AXIS_DOMAIN_CHECK_PATH = os.path.join(os.path.dirname(__file__), 'data', 'Axis_Domain_Check.xlsx')
_DART_NEGATE_CHECK_PATH = os.path.join(os.path.dirname(__file__), 'data', 'DART_Negate_Check.xlsx')


def _load_std():
    """표준 택사노미 로드.
    Axis_Domain_Check.xlsx → DART_Negate_Check.xlsx 순으로 보강.
    """
    std = StandardTaxonomy()

    if os.path.exists(_AXIS_DOMAIN_CHECK_PATH):
        try:
            enrich_axis_domain_check(std, _AXIS_DOMAIN_CHECK_PATH)
        except Exception:
            pass

    if os.path.exists(_DART_NEGATE_CHECK_PATH):
        try:
            enrich_dart_negate_check(std, _DART_NEGATE_CHECK_PATH)
        except Exception:
            pass

    return std


# ─── 직렬화 ──────────────────────────────────────────────────────────────────

def _serialize(data, results, summary) -> dict:
    """CheckResult 목록을 JSON-직렬화 가능한 dict로 변환."""
    out = {
        'company':      data.company_name or '알 수 없음',
        'report_date':  data.report_date  or '',
        'entity_id':    data.entity_id    or '',
        'parse_errors': data.errors,
        'summary': {
            'total_checks':       summary['total_checks'],
            'total_issues':       summary['total_issues'],
            'checks_with_issues': summary['checks_with_issues'],
        },
        'categories':  {},
        'checks_flat': [],
    }

    for cat_name, cat_data in summary['categories'].items():
        out['categories'][cat_name] = {
            'total_issues': cat_data['total_issues'],
            'checks': [],
        }
        for chk in cat_data['checks']:
            issues_out = []
            for iss in chk.issues[:500]:          # 시트당 최대 500건
                iss_dict = {
                    'role_uri':        iss.role_uri,
                    'role_code':       iss.role_code,
                    'role_name_ko':    iss.role_name_ko,
                    'role_name_en':    iss.role_name_en,
                    'is_consolidated': iss.is_consolidated,
                    'element_name':    iss.element_name,
                    'label_ko':        iss.label_ko,
                    'label_en':        iss.label_en,
                    'label_role':      iss.label_role,
                    'prefix':          iss.prefix,
                    'data_type':       iss.data_type,
                    'balance':         iss.balance,
                    'period':          iss.period,
                    'gubn':            iss.gubn,
                    'depth':           iss.depth,
                    'parent_name':     iss.parent_name,
                    'parent_label_ko': iss.parent_label_ko,
                    'parent_gubn':     iss.parent_gubn,
                    'reason':          iss.reason,
                    'table_name_ko':   iss.table_name_ko,
                    'client_negate':   iss.client_negate,
                    'dart_negate':     iss.dart_negate,
                }
                issues_out.append(iss_dict)

            chk_dict = {
                'id':           chk.check_id,
                'title':        chk.title,
                'description':  chk.description,
                'category':     chk.category,
                'sheet':        chk.sheet,
                'issue_count':  chk.issue_count,
                'consol_count': chk.consol_count,
                'sep_count':    chk.sep_count,
                'passed':       chk.passed,
                'issues':       issues_out,
            }
            out['categories'][cat_name]['checks'].append(chk_dict)
            out['checks_flat'].append({
                'id':          chk.check_id,
                'title':       chk.title,
                'issue_count': chk.issue_count,
                'passed':      chk.passed,
            })

    return out


# ─── 라우트 ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/check', methods=['POST'])
def api_check():
    """구조내려받기 xlsx → 29개 체크리스트 실행"""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({
            'error': '.xlsx 파일만 지원합니다. '
                     'IxD 편집기에서 [구조내려받기]로 내보낸 파일을 업로드해 주세요.'
        }), 400

    try:
        xlsx_bytes = f.read()
        data       = parse_taxonomy_xlsx(xlsx_bytes)
        std        = _load_std()
        results    = run_all_checks(data, std)
        summary    = get_summary(results)
        out        = _serialize(data, results, summary)
        _cache['last'] = out
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/export', methods=['GET'])
def api_export():
    """마지막 체크 결과를 XBRL CoE Checklist_Result 템플릿 형식으로 내보내기"""
    cached = _cache.get('last')
    if not cached:
        return jsonify({'error': '먼저 파일을 업로드하세요.'}), 400

    try:
        import openpyxl

        TEMPLATE_PATH = os.path.join(os.path.dirname(__file__),
                                     'template', 'XBRL_CoE_Checklist_Result.xlsx')
        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({'error': '템플릿 파일을 찾을 수 없습니다: template/XBRL_CoE_Checklist_Result.xlsx'}), 500

        def _role_def(iss: dict) -> str:
            rc = iss.get('role_code', ''); ko = iss.get('role_name_ko', ''); en = iss.get('role_name_en', '')
            if rc and ko:
                return f"[{rc}] {ko} | {en}" if en else f"[{rc}] {ko}"
            return ko or rc

        def _table_name(iss: dict) -> str:
            tko = iss.get('table_name_ko', '')
            if tko:
                return tko
            rc = iss.get('role_code', ''); ko = iss.get('role_name_ko', '')
            if rc and ko:
                return f"[{rc}] {ko}"
            return ko or rc

        def _write_standard_row(ws, ri: int, iss: dict):
            ws.cell(ri, 2, _role_def(iss))
            ws.cell(ri, 3, _table_name(iss))
            ws.cell(ri, 4, iss.get('prefix', ''))
            ws.cell(ri, 5, iss.get('element_name', ''))
            ws.cell(ri, 6, iss.get('label_ko', ''))
            ws.cell(ri, 7, iss.get('label_en', ''))
            ws.cell(ri, 8, iss.get('label_role', ''))
            ws.cell(ri, 9, iss.get('data_type', ''))
            ws.cell(ri, 10, iss.get('period', ''))

        def _find_sheet_xml(template_path: str, sheet_name: str) -> str:
            """템플릿 xlsx에서 시트 이름에 해당하는 worksheet XML 경로 반환."""
            with zipfile.ZipFile(template_path, 'r') as z:
                wb_xml   = z.read('xl/workbook.xml').decode('utf-8')
                rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

            escaped = re.escape(sheet_name)
            m = re.search(rf'<sheet\b[^>]*\bname="{escaped}"[^>]*/>', wb_xml)
            if not m:
                raise KeyError(f'workbook.xml에서 시트 찾기 실패: {sheet_name}')
            rid_m = re.search(r'r:id="([^"]+)"', m.group(0))
            if not rid_m:
                raise KeyError(f'<sheet> 태그에서 r:id 추출 실패: {sheet_name}')
            rid = rid_m.group(1)

            m2 = re.search(rf'<Relationship\b[^>]*\bId="{re.escape(rid)}"[^>]*/>', rels_xml)
            if not m2:
                raise KeyError(f'workbook.xml.rels에서 rId 찾기 실패: {rid}')
            target_m = re.search(r'Target="worksheets/([^"]+)"', m2.group(0))
            if not target_m:
                raise KeyError(f'<Relationship> 태그에서 Target 추출 실패: {rid}')
            return f'xl/worksheets/{target_m.group(1)}'

        # ── Step 1: openpyxl로 데이터 기록 후 메모리 버퍼에 저장 ──────────
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

        modified_sheet_names = set()

        for cat_name, cat_data in cached['categories'].items():
            for chk in cat_data['checks']:
                sheet_name = chk.get('sheet', f"Checklist_{chk['id']}")
                sheet_name = sheet_name[:31]

                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # 14행 이후 기존 데이터 클리어
                if ws.max_row >= 14:
                    for row in ws.iter_rows(min_row=14, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None

                # B11에 건수를 정적 값으로 기록 (수식 캐시 0 표시 방지)
                ws['B11'] = len(chk['issues'])
                modified_sheet_names.add(sheet_name)

                chk_id = chk['id']
                for ri, iss in enumerate(chk['issues'], 14):
                    if chk_id == '5-6':
                        ws.cell(ri, 2, _table_name(iss))
                        ws.cell(ri, 3, '단위미표시')
                    elif chk_id == '7-1':
                        _write_standard_row(ws, ri, iss)
                        ws.cell(ri, 11, iss.get('dart_negate', ''))
                        ws.cell(ri, 12, iss.get('client_negate', ''))
                    else:
                        _write_standard_row(ws, ri, iss)

        mod_buf = io.BytesIO()
        wb.save(mod_buf)
        mod_buf.seek(0)

        # ── Step 2: zipfile로 템플릿 구조 보존하며 최종 파일 구성 ──────────
        # 수정된 시트 XML + sharedStrings + styles → openpyxl 버전 사용
        # drawing, media 등 나머지 → 템플릿 원본 사용

        # 수정된 시트들의 XML 경로 확인
        sheet_xml_map: dict = {}   # sheet_name -> xl/worksheets/sheetN.xml
        for sname in modified_sheet_names:
            try:
                sheet_xml_map[sname] = _find_sheet_xml(TEMPLATE_PATH, sname)
            except KeyError:
                pass

        preserve_from_modified = {'xl/sharedStrings.xml', 'xl/styles.xml'}
        preserve_from_modified.update(sheet_xml_map.values())
        # 수정된 시트의 rels 파일도 openpyxl 버전 사용 (rId 재번호 부여 동기화)
        for xml_path in sheet_xml_map.values():
            parts = xml_path.rsplit('/', 1)
            preserve_from_modified.add(f'{parts[0]}/_rels/{parts[1]}.rels')

        # 템플릿 시트 XML에서 drawing/legacyDrawing 참조 태그 추출 (openpyxl이 제거하므로 복원용)
        # <drawing r:id="..."/>        — DrawingML (도형, 차트)
        # <legacyDrawing r:id="..."/> — VML (코멘트, 일부 도형)
        _DRAWING_PAT = re.compile(rb'<(?:legacyDrawing|drawing)[^>]*/>')
        drawing_tags_map: dict = {}   # xml_path -> [bytes, ...]
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip:
            for xml_path in sheet_xml_map.values():
                try:
                    raw = tmpl_zip.read(xml_path)
                    drawing_tags_map[xml_path] = _DRAWING_PAT.findall(raw)
                except Exception:
                    drawing_tags_map[xml_path] = []

        out_buf = io.BytesIO()
        with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip, \
             zipfile.ZipFile(mod_buf, 'r') as mod_zip:
            mod_names = set(mod_zip.namelist())

            with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
                for info in tmpl_zip.infolist():
                    # calcChain 제거: B11 정적 값 교체로 수식 구조가 바뀌었으므로
                    if info.filename == 'xl/calcChain.xml':
                        continue

                    raw = tmpl_zip.read(info.filename)

                    if info.filename in preserve_from_modified and info.filename in mod_names:
                        raw = mod_zip.read(info.filename)

                        # openpyxl이 drawing/legacyDrawing 참조를 제거한 경우 복원
                        drawing_tags = drawing_tags_map.get(info.filename, [])
                        if drawing_tags and not _DRAWING_PAT.search(raw):
                            inject = b''.join(drawing_tags)
                            ws_tag_end = raw.find(b'>', raw.find(b'<worksheet'))
                            if b'xmlns:r=' not in raw[:ws_tag_end]:
                                ns = b' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                                inject = re.sub(
                                    rb'<((?:legacyDrawing|drawing)) ',
                                    rb'<\1' + ns + rb' ',
                                    inject
                                )
                            raw = raw.replace(b'</worksheet>', inject + b'</worksheet>')

                    # [Content_Types].xml에서 calcChain Override 항목 제거
                    if info.filename == '[Content_Types].xml':
                        raw = re.sub(rb'<Override[^>]*calcChain[^>]*/>', b'', raw)

                    # workbook.xml에 fullCalcOnLoad="1" 추가 → Excel이 열 때 전체 재계산
                    if info.filename == 'xl/workbook.xml':
                        raw = re.sub(
                            rb'<calcPr([^/]*)/>', rb'<calcPr\1 fullCalcOnLoad="1"/>', raw
                        )

                    out_zip.writestr(info, raw)

        out_buf.seek(0)

        company = cached.get('company') or ''
        if not company or company == '알 수 없음':
            company = cached.get('entity_id') or 'XBRL'
        rdate   = (cached.get('report_date') or '').replace('/', '').replace('-', '')
        name_parts = [p for p in [company, rdate] if p]
        fname   = f"XBRL_CoE_Checklist_Result_{'_'.join(name_parts)}.xlsx"
        return send_file(
            out_buf, as_attachment=True,
            download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, port=port)
