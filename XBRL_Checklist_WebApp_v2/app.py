"""
app.py  — XBRL 체크리스트 웹 애플리케이션 (Flask)
입력: IxD 편집기 '구조내려받기' .xlsx 파일 전용 
"""
import io, os, json, traceback
from flask import Flask, request, jsonify, render_template, send_file

from taxonomy_xlsx_parser import parse_taxonomy_xlsx
from checklist_engine import run_all_checks, get_summary

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024   # 100 MB

# 단순 메모리 캐시 (프로세스 내 마지막 결과 보관)
_cache: dict = {}


# ─── 직렬화 ──────────────────────────────────────────────────────────────────

def _serialize(data, results, summary) -> dict:
    """CheckResult 목록을 JSON-직렬화 가능한 dict로 변환."""
    out = {
        'company':      data.company_name or '알 수 없음',
        'report_date':  data.report_date  or '',
        'entity_id':    data.entity_id    or '',
        'parse_errors': data.errors,
        'summary': {
            'total_checks':       summary['total_checks'],
            'total_issues':       summary['total_issues'],
            'checks_with_issues': summary['checks_with_issues'],
        },
        'categories':  {},
        'checks_flat': [],
    }

    for cat_name, cat_data in summary['categories'].items():
        out['categories'][cat_name] = {
            'total_issues': cat_data['total_issues'],
            'checks': [],
        }
        for chk in cat_data['checks']:
            issues_out = []
            for iss in chk.issues[:500]:          # 시트당 최대 500건
                iss_dict = {
                    'role_uri':        iss.role_uri,
                    'role_code':       iss.role_code,
                    'role_name_ko':    iss.role_name_ko,
                    'role_name_en':    iss.role_name_en,
                    'is_consolidated': iss.is_consolidated,
                    'element_name':    iss.element_name,
                    'label_ko':        iss.label_ko,
                    'label_en':        iss.label_en,
                    'prefix':          iss.prefix,
                    'data_type':       iss.data_type,
                    'balance':         iss.balance,
                    'period':          iss.period,
                    'gubn':            iss.gubn,
                    'depth':           iss.depth,
                    'parent_name':     iss.parent_name,
                    'parent_label_ko': iss.parent_label_ko,
                    'parent_gubn':     iss.parent_gubn,
                    'reason':          iss.reason,
                    'table_name_ko':   iss.table_name_ko,
                }
                if chk.check_id == '7-1':
                    iss_dict['client_negate'] = iss.client_negate
                    iss_dict['dart_negate']   = iss.dart_negate
                issues_out.append(iss_dict)

            chk_dict = {
                'id':           chk.check_id,
                'title':        chk.title,
                'description':  chk.description,
                'category':     chk.category,
                'sheet':        chk.sheet,
                'issue_count':  chk.issue_count,
                'consol_count': chk.consol_count,
                'sep_count':    chk.sep_count,
                'passed':       chk.passed,
                'issues':       issues_out,
            }
            out['categories'][cat_name]['checks'].append(chk_dict)
            out['checks_flat'].append({
                'id':          chk.check_id,
                'title':       chk.title,
                'issue_count': chk.issue_count,
                'passed':      chk.passed,
            })

    return out


# ─── 라우트 ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/check', methods=['POST'])
def api_check():
    """구조내려받기 xlsx → 29개 체크리스트 실행"""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400

    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({
            'error': '.xlsx 파일만 지원합니다. '
                     'IxD 편집기에서 [구조내려받기]로 내보낸 파일을 업로드해 주세요.'
        }), 400

    try:
        xlsx_bytes = f.read()
        data       = parse_taxonomy_xlsx(xlsx_bytes)
        results    = run_all_checks(data)
        summary    = get_summary(results)
        out        = _serialize(data, results, summary)
        _cache['last'] = out
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


@app.route('/api/export', methods=['GET'])
def api_export():
    """마지막 체크 결과를 Excel 파일로 내보내기"""
    cached = _cache.get('last')
    if not cached:
        return jsonify({'error': '먼저 파일을 업로드하세요.'}), 400

    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── 스타일 상수 ────────────────────────────────────────────────────
        HDR_FILL  = PatternFill('solid', fgColor='0D2345')
        HDR_FONT  = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=9)
        ERR_FILL  = PatternFill('solid', fgColor='FFF2F2')
        OK_FILL   = PatternFill('solid', fgColor='F0FFF4')
        WARN_FILL = PatternFill('solid', fgColor='FFFBE6')
        THIN      = Side(style='thin', color='D1D5DB')
        BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        CTR       = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        DATA_FONT = Font(name='맑은 고딕', size=9)

        def hdr(ws, row_idx, cols):
            for ci, (title, width) in enumerate(cols, 1):
                c = ws.cell(row_idx, ci, title)
                c.font = HDR_FONT; c.fill = HDR_FILL
                c.alignment = CTR; c.border = BORDER
                ws.column_dimensions[get_column_letter(ci)].width = width
            ws.row_dimensions[row_idx].height = 22

        def cell(ws, r, c, v, fill=None, align=None):
            o = ws.cell(r, c, v)
            o.font  = DATA_FONT
            o.border = BORDER
            o.alignment = align or LEFT
            if fill: o.fill = fill
            return o

        # ── 요약 시트 ──────────────────────────────────────────────────────
        ws0 = wb.create_sheet('요약')
        ws0.freeze_panes = 'A2'

        # 상단 정보
        info_rows = [
            ('회사명',   cached.get('company', '')),
            ('보고일',   cached.get('report_date', '')),
            ('총 체크',  cached['summary']['total_checks']),
            ('총 검출',  cached['summary']['total_issues']),
            ('이상 항목',cached['summary']['checks_with_issues']),
        ]
        for ri, (k, v) in enumerate(info_rows, 1):
            ws0.cell(ri, 1, k).font = Font(bold=True, name='맑은 고딕', size=9)
            ws0.cell(ri, 2, v).font = DATA_FONT
        ws0.column_dimensions['A'].width = 14
        ws0.column_dimensions['B'].width = 30

        START_ROW = len(info_rows) + 2

        sum_cols = [
            ('체크 ID', 9), ('체크 항목명', 34), ('카테고리', 22),
            ('시트명', 18), ('검출(전체)', 10), ('연결', 7), ('별도', 7), ('결과', 8),
        ]
        hdr(ws0, START_ROW, sum_cols)

        r = START_ROW + 1
        for cat_name, cat_data in cached['categories'].items():
            for chk in cat_data['checks']:
                fl = OK_FILL if chk['passed'] else ERR_FILL
                cell(ws0, r, 1, chk['id'],           fl, CTR)
                cell(ws0, r, 2, chk['title'],        fl)
                cell(ws0, r, 3, cat_name,            fl)
                cell(ws0, r, 4, chk.get('sheet',''), fl)
                cell(ws0, r, 5, chk['issue_count'],  fl, CTR)
                cell(ws0, r, 6, chk['consol_count'], fl, CTR)
                cell(ws0, r, 7, chk['sep_count'],    fl, CTR)
                result_cell = cell(ws0, r, 8,
                                   '통과' if chk['passed'] else '검출', fl, CTR)
                if not chk['passed']:
                    result_cell.font = Font(name='맑은 고딕', size=9,
                                            color='C00000', bold=True)
                r += 1

        # ── 체크별 상세 시트 ────────────────────────────────────────────────
        detail_cols = [
            ('연결/별도',  8), ('Role 코드',  13), ('Role 한글명', 26), ('Role 영문명', 36),
            ('요소명(Name)', 36), ('요소 한글명', 22), ('요소 영문명', 30),
            ('Prefix', 11), ('DataType', 20), ('Balance', 8), ('Period', 9),
            ('구분', 10), ('상위 요소', 22), ('검토 사유', 35),
        ]

        for cat_name, cat_data in cached['categories'].items():
            for chk in cat_data['checks']:
                sname = chk.get('sheet', f"Checklist_{chk['id']}")
                # 시트 이름 31자 제한
                ws = wb.create_sheet(sname[:31])
                ws.freeze_panes = 'A2'
                hdr(ws, 1, detail_cols)

                if not chk['issues']:
                    c = ws.cell(2, 1, '검출된 항목이 없습니다.')
                    c.font = Font(name='맑은 고딕', size=9,
                                  color='16A34A', bold=True)
                    c.alignment = CTR
                    ws.merge_cells(
                        f'A2:{get_column_letter(len(detail_cols))}2')
                    ws.row_dimensions[2].height = 18
                    continue

                for ri, iss in enumerate(chk['issues'], 2):
                    cs = ('연결' if iss['is_consolidated'] is True
                          else '별도' if iss['is_consolidated'] is False
                          else '-')
                    cell(ws, ri, 1,  cs,                   ERR_FILL, CTR)
                    cell(ws, ri, 2,  iss['role_code'],      ERR_FILL, CTR)
                    cell(ws, ri, 3,  iss['role_name_ko'],   ERR_FILL)
                    cell(ws, ri, 4,  iss['role_name_en'],   ERR_FILL)
                    cell(ws, ri, 5,  iss['element_name'],   ERR_FILL)
                    cell(ws, ri, 6,  iss['label_ko'],       ERR_FILL)
                    cell(ws, ri, 7,  iss['label_en'],       ERR_FILL)
                    cell(ws, ri, 8,  iss['prefix'],         ERR_FILL, CTR)
                    cell(ws, ri, 9,  iss['data_type'],      ERR_FILL)
                    cell(ws, ri, 10, iss['balance'],        ERR_FILL, CTR)
                    cell(ws, ri, 11, iss['period'],         ERR_FILL, CTR)
                    cell(ws, ri, 12, iss['gubn'],           ERR_FILL, CTR)
                    cell(ws, ri, 13, iss['parent_label_ko'],ERR_FILL)
                    cell(ws, ri, 14, iss['reason'],         ERR_FILL)
                    ws.row_dimensions[ri].height = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        company = cached.get('company', 'XBRL')
        rdate   = (cached.get('report_date') or '').replace('/', '').replace('-', '')
        fname   = f"XBRL_Checklist_{company}_{rdate}.xlsx"
        return send_file(
            buf, as_attachment=True,
            download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, port=port)
    
    